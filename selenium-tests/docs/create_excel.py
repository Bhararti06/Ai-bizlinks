import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Read the CSV file
df = pd.read_csv('docs/Test_Cases_Data.csv')

# Create Excel writer
excel_file = 'docs/Test_Cases_Data.xlsx'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# Write to Excel
df.to_excel(writer, sheet_name='Test Cases', index=False)

# Get the workbook and worksheet
workbook = writer.book
worksheet = writer.sheets['Test Cases']

# Style the header row
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Auto-adjust column widths
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
    worksheet.column_dimensions[column_letter].width = adjusted_width

# Apply conditional formatting for Status column (column C)
for row in range(2, len(df) + 2):
    status_cell = worksheet[f'C{row}']
    if status_cell.value == 'PASS':
        status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        status_cell.font = Font(color='006100', bold=True)
    elif status_cell.value == 'FAIL':
        status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        status_cell.font = Font(color='9C0006', bold=True)

# Apply conditional formatting for Priority column (column D)
for row in range(2, len(df) + 2):
    priority_cell = worksheet[f'D{row}']
    if priority_cell.value == 'CRITICAL':
        priority_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        priority_cell.font = Font(color='FFFFFF', bold=True)
    elif priority_cell.value == 'HIGH':
        priority_cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
        priority_cell.font = Font(bold=True)
    elif priority_cell.value == 'MEDIUM':
        priority_cell.fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
        priority_cell.font = Font(color='FFFFFF')

# Freeze the header row
worksheet.freeze_panes = 'A2'

# Save the workbook
writer.close()

print(f"✅ Excel file created successfully: {excel_file}")
print(f"📊 Total test cases: {len(df)}")
print(f"✅ Passing: {len(df[df['Status'] == 'PASS'])}")
print(f"❌ Failing: {len(df[df['Status'] == 'FAIL'])}")
